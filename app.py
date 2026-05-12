"""
爆单猫-TikTok利润测算系统 — Streamlit云端版
功能：通过浏览器上传Excel，在线完成利润测算并下载报告
"""

import streamlit as st
import os
import sys
import tempfile
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO

# ── 页面配置 ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="爆单猫-TikTok利润测算",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS样式 ───────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-title { font-size: 2rem; font-weight: 700; color: #1f1f1f; }
    .subtitle { font-size: 1rem; color: #666; margin-bottom: 2rem; }
    .metric-card { background: #f8f9fa; border-radius: 10px; padding: 1rem; text-align: center; }
    .metric-value { font-size: 1.8rem; font-weight: 700; color: #0068c9; }
    .metric-label { font-size: 0.9rem; color: #666; }
    .upload-section { background: #fafafa; border: 1px dashed #ccc; border-radius: 8px; padding: 1rem; }
    div[data-testid="stFileUploader"] { width: 100%; }
</style>
""", unsafe_allow_html=True)

# ── 路径设置 ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SYS_DIR = os.path.join(BASE_DIR, '00-系统文件（不动）')
sys.path.insert(0, SYS_DIR)

# ── 导入核心模块 ──────────────────────────────────────────────────────────────
from data_loader_v3 import DataLoader
from calculator_v3 import ProfitCalculator
from validation import generate_validation_sheets
from baseline_loader import load_last_year_orders

# ── Session State ─────────────────────────────────────────────────────────────
if 'calculated' not in st.session_state:
    st.session_state.calculated = False
if 'result_excel' not in st.session_state:
    st.session_state.result_excel = None
if 'report_period' not in st.session_state:
    st.session_state.report_period = ''


# ═══════════════════════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════════════════════

def calculate_report_period(data) -> str:
    """计算店铺订单的数据周期"""
    MIN_DATE = datetime(datetime.now().year - 1, 1, 1)
    MAX_DATE = datetime.now() + timedelta(days=30)
    if 'orders' in data and not data['orders'].empty:
        df = data['orders']
        if 'Created Time' in df.columns:
            try:
                s = pd.to_datetime(df['Created Time'], errors='coerce')
                s = s[(s >= MIN_DATE) & (s <= MAX_DATE)]
                if not s.empty:
                    return f"{s.min().strftime('%Y-%m-%d')} ~ {s.max().strftime('%Y-%m-%d')}"
            except Exception:
                pass
    return ''


def save_uploaded_file(uploaded_file, save_path):
    """保存上传的文件到指定路径"""
    if uploaded_file is not None:
        with open(save_path, 'wb') as f:
            f.write(uploaded_file.getvalue())
        return True
    return False


def build_custom_paths(temp_dir, uploaded_files):
    """构建custom_paths字典"""
    base_data = os.path.join(temp_dir, '01-基础数据（年度更新）')
    current_data = os.path.join(temp_dir, '02-本期数据（每次测算更新）')
    os.makedirs(base_data, exist_ok=True)
    os.makedirs(current_data, exist_ok=True)

    paths = {}
    file_mapping = {
        # 基础数据
        'spu_map': (base_data, 'spu_sku映射表.xlsx'),
        'sku_category_map': (base_data, 'SKU-三级类目映射.xlsx'),
        'pid_sku_map': (base_data, 'pid_sku映射表.xlsx'),
        # 本期数据
        'orders': (current_data, '店铺订单.xlsx'),
        'purchase': (current_data, '采购成本.xlsx'),
        'first_mile': (current_data, '采购成本.xlsx'),
        'last_mile': (current_data, '尾程成本.xlsx'),
        'tariff': (current_data, '关税成本.xlsx'),
        'affiliate': (current_data, '联盟订单.xlsx'),
        'ads': (current_data, '广告订单.xlsx'),
        'excluded_orders': (current_data, '需排除的订单编号.xlsx'),
    }

    for key, (folder, filename) in file_mapping.items():
        if key in uploaded_files and uploaded_files[key] is not None:
            path = os.path.join(folder, filename)
            save_uploaded_file(uploaded_files[key], path)
            paths[key] = path

    return paths, base_data, current_data


def read_base_params(param_file):
    """从上传的基础参数表读取参数"""
    defaults = {'exchange_rate': 7.1, 'platform_commission_rate': 0.09, 'other_fee_rate': 0.015}
    if param_file is None:
        return defaults
    try:
        df = pd.read_excel(param_file, sheet_name='基础参数')
        for _, row in df.iterrows():
            name = str(row.get('参数名称', '')).strip()
            val = row.get('参数值', None)
            if pd.isna(val):
                continue
            if name == '汇率':
                defaults['exchange_rate'] = float(val)
            elif name == '平台佣金率':
                defaults['platform_commission_rate'] = float(val)
            elif name == '其他费用率':
                defaults['other_fee_rate'] = float(val)
    except Exception:
        pass
    return defaults


def generate_excel_report(result, validation, report_period, params_info):
    """生成Excel报告并返回bytes"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output = BytesIO()

    # 广告费分摊明细处理
    spu_ads = result.cost_allocation
    if '三级类目' in result.spu_profit.columns and not spu_ads.empty:
        category_map = result.spu_profit[['SPU', '三级类目']].drop_duplicates()
        spu_ads = spu_ads.merge(category_map, on='SPU', how='left')
        cols = ['SPU', '三级类目'] + [c for c in spu_ads.columns if c not in ['SPU', '三级类目']]
        spu_ads = spu_ads[cols]

    # 数据说明
    df_info = pd.DataFrame({
        '项目': [
            '测算周期', '生成时间', '汇率', '平台佣金率', '其他费用率',
            '店铺订单文件', '广告订单文件', '联盟订单文件',
            '采购成本文件', '尾程成本文件', '关税成本文件'
        ],
        '内容': [
            report_period or '-',
            timestamp,
            params_info.get('exchange_rate', 7.1),
            f"{params_info.get('platform_commission_rate', 0.09) * 100:.1f}%",
            f"{params_info.get('other_fee_rate', 0.015) * 100:.1f}%",
            '店铺订单.xlsx', '广告订单.xlsx', '联盟订单.xlsx',
            '采购成本.xlsx', '尾程成本.xlsx', '关税成本.xlsx'
        ]
    })

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result.store_matrix.to_excel(writer, sheet_name='店铺利润汇总', index=False)
        result.spu_matrix.to_excel(writer, sheet_name='SPU利润汇总', index=False)
        result.category_matrix.to_excel(writer, sheet_name='三级类目利润汇总', index=False)
        spu_ads.to_excel(writer, sheet_name='广告费分摊明细', index=False)
        df_info.to_excel(writer, sheet_name='数据说明', index=False)
        validation['sku_cost'].to_excel(writer, sheet_name='数据校验', index=False)

    # openpyxl后处理：标黄缺失项
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    output.seek(0)
    wb = load_workbook(output)
    ws = wb['数据校验']

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    df_sku = validation['sku_cost']
    cost_cols = ['有采购成', '有头程成', '有尾程成', '有关税成']
    cost_col_indices = [df_sku.columns.get_loc(c) for c in cost_cols]

    for excel_row, row_values in enumerate(df_sku.values, start=2):
        for col_idx in cost_col_indices:
            if row_values[col_idx] != '√':
                ws.cell(row=excel_row, column=col_idx + 1).fill = yellow_fill

    def append_df(ws, df, start_row, title=None):
        if title:
            cell = ws.cell(row=start_row, column=1, value=title)
            cell.font = Font(bold=True)
            start_row += 1
        if df.empty:
            return start_row + 2
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=start_row, column=col_idx, value=col_name)
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        return start_row + len(df) + 1

    current_row = ws.max_row + 3
    current_row = append_df(ws, validation['affiliate'], current_row)
    current_row += 2
    current_row = append_df(ws, validation['ads'], current_row)
    current_row += 2
    current_row = append_df(ws, validation['ads_unmapped'], current_row, title='广告费未匹配明细')
    current_row += 2
    append_df(ws, validation['other'], current_row)

    # 保存到BytesIO
    output.seek(0)
    output.truncate(0)
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.getvalue(), timestamp


# ═══════════════════════════════════════════════════════════════════════════════
# 渲染辅助函数
# ═══════════════════════════════════════════════════════════════════════════════

def extract_value_rows(df, id_col):
    """从2行/3行矩阵中提取数值行（id_col非空的行为数值行）"""
    if id_col in df.columns:
        mask = df[id_col].notna() & (df[id_col].astype(str).str.strip() != '')
        return df[mask].reset_index(drop=True)
    return df.iloc[[0]].reset_index(drop=True)


def render_detail_table(df):
    """渲染紧凑详情表格（所有列在一屏内显示，无滚动条）"""
    cols = df.columns.tolist()
    header_cells = ''.join(
        f'<th style="padding:2px 3px; font-size:9px; text-align:center; white-space:nowrap; '
        f'background:#2c3e50; color:white; font-weight:500; border:1px solid #3d566e;">{c}</th>'
        for c in cols
    )
    body_rows = ''
    for i, row in df.iterrows():
        cells = ''.join(
            f'<td style="padding:2px 3px; font-size:9px; text-align:center; white-space:nowrap; '
            f"border-bottom:1px solid #f0f0f0; {'background:#fafafa;' if i % 2 == 1 else ''}\">{v}</td>"
            for v in row
        )
        body_rows += f'<tr>{cells}</tr>'
    return f'<table style="width:100%; border-collapse:collapse;">{header_cells}{body_rows}</table>'


def render_store_table(df):
    """渲染店铺利润汇总HTML表格（参考图片排版）"""
    cols = df.columns.tolist()
    # 表头
    header_cells = ''.join(
        f'<th style="padding:10px 8px; text-align:center; font-weight:600; font-size:13px; white-space:nowrap;">{c}</th>'
        for c in cols
    )
    # 数据行
    body_rows = []
    for i, row in df.iterrows():
        cells = []
        for col in cols:
            val = row.get(col, '')
            if i == 0:
                # 数值行：粗体，金额加$
                style = 'padding:10px 8px; text-align:center; font-weight:600; font-size:14px;'
                display = val
            elif i == 1:
                # 同比行：绿色小字
                style = 'padding:6px 8px; text-align:center; color:#4caf50; font-size:12px;'
                display = val
            else:
                # 占比行：灰色小字
                style = 'padding:6px 8px; text-align:center; color:#999; font-size:12px;'
                display = val
            # 毛利率和毛利润高亮
            if col == '毛利率' and i == 2 and str(val).replace('%', '').replace('-', '').replace('.', '').strip():
                try:
                    v = float(str(val).replace('%', ''))
                    color = '#4caf50' if v > 0 else '#f44336'
                    style = style.replace('color:#999', f'color:{color}')
                    style += ' font-weight:700;'
                except:
                    pass
            cells.append(f'<td style="{style}">{display}</td>')
        body_rows.append('<tr style="border-bottom:1px solid #f0f0f0;">' + ''.join(cells) + '</tr>')

    html = f'''
    <table style="width:100%; border-collapse:collapse; font-size:14px; margin:10px 0;">
        <thead>
            <tr style="background:#2c3e50; color:white;">{header_cells}</tr>
        </thead>
        <tbody>
            {''.join(body_rows)}
        </tbody>
    </table>
    '''
    return html


def render_item_cards(df, id_col, title):
    """渲染SPU/类目卡片列表（含筛选排序交互控件）"""
    # 提取数值行和占比行
    value_rows = extract_value_rows(df, id_col)
    pct_rows = []
    for idx in value_rows.index:
        # 占比行在数值行之后
        if idx + 1 < len(df):
            pct_rows.append(df.iloc[idx + 1])
        else:
            pct_rows.append(pd.Series())

    # 转为DataFrame便于操作
    records = []
    for i, (vrow, prow) in enumerate(zip(value_rows.itertuples(index=False), pct_rows)):
        vdict = vrow._asdict()
        vdict['_pct_row'] = prow
        vdict['_orig_idx'] = value_rows.index[i]
        records.append(vdict)

    if not records:
        st.info(f"暂无{title}数据")
        return

    rec_df = pd.DataFrame(records)

    # 确保数值类型正确
    numeric_cols = ['退款前营收', '退款后营收', '退款前销量', '退款后销量', '采购成本', '头程成本',
                    '尾程成本', '关税成本', '联盟佣金', '广告费', '平台佣金', '其他费用', '样品费',
                    '退款金额', '毛利润']
    for c in numeric_cols:
        if c in rec_df.columns:
            rec_df[c] = pd.to_numeric(rec_df[c], errors='coerce').fillna(0)

    # 计算利润率用于筛选
    if '毛利率' in rec_df.columns:
        rec_df['_margin_num'] = rec_df['毛利率'].apply(lambda x: float(str(x).replace('%', '')) if pd.notna(x) and str(x) != '—' else 0)
    else:
        rec_df['_margin_num'] = 0

    # ── 筛选 & 排序控件 ─────────────────────────────────────────
    ctrl1, ctrl2, ctrl3 = st.columns([2, 2, 3])

    with ctrl1:
        filter_type = st.segmented_control(
            "筛选", ["全部", "盈利", "亏损"],
            default="全部", key=f"{title}_filter"
        )
    with ctrl2:
        sort_by = st.segmented_control(
            "排序", ["按营收", "按销量"],
            default="按营收", key=f"{title}_sort"
        )

    total = len(rec_df)
    profit_count = len(rec_df[rec_df['毛利润'] > 0])
    loss_count = len(rec_df[rec_df['毛利润'] < 0])

    with ctrl3:
        st.markdown(f"<div style='padding-top:28px; color:#666; font-size:13px;'>共 {total} 个 | 盈利 {profit_count} | 亏损 {loss_count}</div>", unsafe_allow_html=True)

    # 筛选
    if filter_type == "盈利":
        rec_df = rec_df[rec_df['毛利润'] > 0]
    elif filter_type == "亏损":
        rec_df = rec_df[rec_df['毛利润'] < 0]

    # 排序
    if sort_by == "按营收":
        rec_df = rec_df.sort_values('退款前营收', ascending=False)
    else:
        rec_df = rec_df.sort_values('退款后销量', ascending=False)

    # ── 卡片列表 ─────────────────────────────────────────────────
    for _, row in rec_df.iterrows():
        name = row.get(id_col, '')
        cat = row.get('三级类目', '')
        profit = row.get('毛利润', 0)
        margin_str = str(row.get('毛利率', '—'))
        revenue = row.get('退款前营收', 0)
        net_revenue = row.get('退款后营收', 0)
        pct_row = row.get('_pct_row', pd.Series())
        revenue_pct = str(pct_row.get('营收占比', '—')) if hasattr(pct_row, 'get') else '—'

        # 计算衍生指标
        ad_spend = row.get('广告费', 0)
        refund_amt = row.get('退款金额', 0)
        ad_ratio = ad_spend / net_revenue * 100 if net_revenue > 0 else 0
        refund_ratio = refund_amt / revenue * 100 if revenue > 0 else 0

        # 标签 & 颜色
        is_profit = profit > 0
        tag = "⭐ 主力产品" if is_profit else "⚠️ 亏损主力"
        tag_bg = "#e3f2fd" if is_profit else "#ffebee"
        tag_color = "#1976d2" if is_profit else "#d32f2f"
        bar_color = "#4caf50" if is_profit else "#f44336"

        # 利润率颜色
        margin_val = row.get('_margin_num', 0)
        margin_color = "#4caf50" if margin_val > 0 else "#f44336"

        # 利润颜色
        profit_color = "#4caf50" if profit > 0 else "#f44336"
        profit_str = f"${profit:,.0f}" if profit >= 0 else f"-${abs(profit):,.0f}"

        # 卡片HTML
        card_html = f'''
        <div style="border-left:4px solid {bar_color}; background:#fafafa; padding:12px 16px;
                    margin:6px 0; border-radius:0 8px 8px 0;">
            <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:8px;">
                <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
                    <span style="background:{tag_bg}; color:{tag_color}; padding:2px 8px;
                                border-radius:4px; font-size:12px; font-weight:500;">{tag}</span>
                    <span style="font-weight:700; font-size:15px; color:#1f1f1f;">{name}</span>
                    <span style="color:#666; font-size:13px;">{cat}</span>
                </div>
                <div style="display:flex; align-items:center; gap:14px; font-size:13px; flex-wrap:wrap;">
                    <span style="background:#e3f2fd; color:#1976d2; padding:2px 10px;
                                border-radius:12px; font-weight:500;">营收占比 {revenue_pct}</span>
                    <span style="color:#888;">广告占比 {ad_ratio:.1f}%</span>
                    <span style="color:#888;">退款率 {refund_ratio:.1f}%</span>
                    <span style="color:{margin_color}; font-weight:700;">{margin_str}</span>
                    <span style="color:{profit_color}; font-weight:700; font-size:15px;">{profit_str}</span>
                </div>
            </div>
        </div>
        '''
        st.markdown(card_html, unsafe_allow_html=True)

        # ── 展开详情：完整矩阵数据 ───────────────────────────────
        orig_idx = row.get('_orig_idx', 0)
        detail_df = df.iloc[orig_idx:orig_idx+2].reset_index(drop=True)
        with st.expander("📋 查看完整数据"):
            st.markdown(render_detail_table(detail_df), unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# 页面主体
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown('<div class="main-title">📊 爆单猫-TikTok利润测算系统</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">上传数据文件，一键生成利润分析报告</div>', unsafe_allow_html=True)

# ── 侧边栏：参数设置 ─────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 基础参数")
    st.markdown("参数可从「基础参数表」自动读取，也可手动覆盖。")

    col1, col2 = st.columns(2)
    with col1:
        exchange_rate = st.number_input("汇率", value=7.1, step=0.1, format="%.2f", key="param_rate")
    with col2:
        platform_commission = st.number_input("平台佣金率", value=0.09, step=0.01, format="%.3f", key="param_commission")
    other_fee = st.number_input("其他费用率", value=0.015, step=0.005, format="%.3f", key="param_other")

    st.divider()
    st.info("💡 提示：修改上方参数后，需要重新点击「开始测算」生效。")

# ── 文件上传区域 ─────────────────────────────────────────────────────────────
st.subheader("📁 数据文件上传")

# 上传模式切换
upload_mode = st.radio(
    "上传方式",
    ["批量上传（推荐）", "单独上传"],
    horizontal=True,
    help="批量上传：打开文件夹，按住 Cmd+A（Mac）或 Ctrl+A（Windows）全选所有 Excel 文件，一次性上传"
)

# 初始化所有文件变量为None
spu_map_file = sku_cat_file = pid_sku_file = param_file = last_year_file = None
orders_file = purchase_file = lastmile_file = tariff_file = affiliate_file = ads_file = excluded_file = None

if upload_mode == "批量上传（推荐）":
    batch_files = st.file_uploader(
        "上传所有 Excel 文件（支持多选）",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        key='batch_upload'
    )

    # 文件名到变量的映射（去掉扩展名匹配）
    filename_map = {
        'spu_sku映射表': 'spu_map',
        'SKU-三级类目映射': 'sku_category',
        'pid_sku映射表': 'pid_sku',
        '基础参数表': 'param',
        '2025年全年订单底表': 'last_year',
        '店铺订单': 'orders',
        '采购成本': 'purchase',
        '尾程成本': 'lastmile',
        '关税成本': 'tariff',
        '联盟订单': 'affiliate',
        '广告订单': 'ads',
        '需排除的订单编号': 'excluded',
    }

    recognized = {}
    unrecognized = []

    if batch_files:
        for f in batch_files:
            name_no_ext = os.path.splitext(f.name)[0]
            if name_no_ext in filename_map:
                recognized[name_no_ext] = f
            else:
                # 尝试模糊匹配（去掉空格、特殊字符）
                normalized = name_no_ext.replace(' ', '').replace('-', '').replace('_', '').lower()
                matched = False
                for key in filename_map:
                    if normalized == key.replace(' ', '').replace('-', '').replace('_', '').lower():
                        recognized[key] = f
                        matched = True
                        break
                if not matched:
                    unrecognized.append(f.name)

        # 赋值给变量
        spu_map_file = recognized.get('spu_sku映射表')
        sku_cat_file = recognized.get('SKU-三级类目映射')
        pid_sku_file = recognized.get('pid_sku映射表')
        param_file = recognized.get('基础参数表')
        last_year_file = recognized.get('2025年全年订单底表')
        orders_file = recognized.get('店铺订单')
        purchase_file = recognized.get('采购成本')
        lastmile_file = recognized.get('尾程成本')
        tariff_file = recognized.get('关税成本')
        affiliate_file = recognized.get('联盟订单')
        ads_file = recognized.get('广告订单')
        excluded_file = recognized.get('需排除的订单编号')

        # 显示识别结果
        if recognized:
            with st.expander(f"✅ 已识别 {len(recognized)} 个文件", expanded=True):
                cols = st.columns(2)
                base_files = [k for k in recognized if k in ['spu_sku映射表', 'SKU-三级类目映射', 'pid_sku映射表', '基础参数表', '2025年全年订单底表']]
                current_files = [k for k in recognized if k not in base_files]
                with cols[0]:
                    st.markdown("**基础数据**")
                    for name in base_files:
                        st.markdown(f"<span style='color:green'>✓</span> {name}.xlsx", unsafe_allow_html=True)
                with cols[1]:
                    st.markdown("**本期数据**")
                    for name in current_files:
                        st.markdown(f"<span style='color:green'>✓</span> {name}.xlsx", unsafe_allow_html=True)

        if unrecognized:
            st.warning(f"⚠️ 以下文件无法自动识别，请检查文件名：{', '.join(unrecognized)}")

else:
    # 单独上传模式
    upload_col1, upload_col2 = st.columns(2)

    with upload_col1:
        st.markdown("**基础数据（年度更新）**")
        spu_map_file = st.file_uploader("SPU-SKU映射表 *", type=['xlsx', 'xls'], key='spu_map')
        sku_cat_file = st.file_uploader("SKU-三级类目映射 *", type=['xlsx', 'xls'], key='sku_cat')
        pid_sku_file = st.file_uploader("PID-SKU映射表", type=['xlsx', 'xls'], key='pid_sku')
        param_file = st.file_uploader("基础参数表 *", type=['xlsx', 'xls'], key='param')
        last_year_file = st.file_uploader("2025年全年订单底表（用于同比）", type=['xlsx', 'xls'], key='last_year')

    with upload_col2:
        st.markdown("**本期数据（每次测算更新）**")
        orders_file = st.file_uploader("店铺订单 *", type=['xlsx', 'xls'], key='orders')
        purchase_file = st.file_uploader("采购成本 *", type=['xlsx', 'xls'], key='purchase')
        lastmile_file = st.file_uploader("尾程成本 *", type=['xlsx', 'xls'], key='lastmile')
        tariff_file = st.file_uploader("关税成本 *", type=['xlsx', 'xls'], key='tariff')
        affiliate_file = st.file_uploader("联盟订单 *", type=['xlsx', 'xls'], key='affiliate')
        ads_file = st.file_uploader("广告订单 *", type=['xlsx', 'xls'], key='ads')
        excluded_file = st.file_uploader("需排除的订单编号", type=['xlsx', 'xls'], key='excluded')

# 自动读取上传的参数表并更新侧边栏
if param_file is not None:
    params_from_file = read_base_params(param_file)
    if st.session_state.param_rate == 7.1 and params_from_file['exchange_rate'] != 7.1:
        st.session_state.param_rate = params_from_file['exchange_rate']
    if st.session_state.param_commission == 0.09 and params_from_file['platform_commission_rate'] != 0.09:
        st.session_state.param_commission = params_from_file['platform_commission_rate']
    if st.session_state.param_other == 0.015 and params_from_file['other_fee_rate'] != 0.015:
        st.session_state.param_other = params_from_file['other_fee_rate']

# ── 检查必填文件 ─────────────────────────────────────────────────────────────
required_files = {
    'spu_map': spu_map_file, 'sku_category_map': sku_cat_file,
    'orders': orders_file, 'purchase': purchase_file,
    'last_mile': lastmile_file, 'tariff': tariff_file,
    'affiliate': affiliate_file, 'ads': ads_file,
}
all_ready = all(f is not None for f in required_files.values())

# ── 开始测算按钮 ─────────────────────────────────────────────────────────────
btn_col1, btn_col2, btn_col3 = st.columns([1, 2, 1])
with btn_col2:
    start_btn = st.button("🚀 开始测算", type="primary", disabled=not all_ready, use_container_width=True)

if not all_ready:
    missing_names = {
        'spu_map': 'SPU-SKU映射表', 'sku_category_map': 'SKU-三级类目映射',
        'orders': '店铺订单', 'purchase': '采购成本',
        'last_mile': '尾程成本', 'tariff': '关税成本',
        'affiliate': '联盟订单', 'ads': '广告订单',
    }
    missing = [missing_names.get(k, k) for k, v in required_files.items() if v is None]
    st.warning(f"⏳ 请上传以下必填文件后再开始测算：{', '.join(missing)}")

# ═══════════════════════════════════════════════════════════════════════════════
# 测算执行
# ═══════════════════════════════════════════════════════════════════════════════

if start_btn and all_ready:
    progress_bar = st.progress(0, text="准备中...")

    try:
        # Step 1: 创建临时目录并保存文件
        progress_bar.progress(5, text="📁 保存上传文件...")
        with tempfile.TemporaryDirectory() as tmpdir:
            uploaded_files = {
                'spu_map': spu_map_file, 'sku_category_map': sku_cat_file,
                'pid_sku_map': pid_sku_file,
                'orders': orders_file, 'purchase': purchase_file,
                'first_mile': purchase_file,  # 头程成本与采购成本在同一文件
                'last_mile': lastmile_file, 'tariff': tariff_file,
                'affiliate': affiliate_file, 'ads': ads_file,
                'excluded_orders': excluded_file,
            }
            custom_paths, base_data_dir, current_data_dir = build_custom_paths(tmpdir, uploaded_files)

            # 保存基础参数表（覆盖参数值）
            if param_file is not None:
                param_path = os.path.join(base_data_dir, '基础参数表.xlsx')
                save_uploaded_file(param_file, param_path)
                # 覆盖参数值
                df_param = pd.read_excel(param_path)
                for _, row in df_param.iterrows():
                    name = str(row.get('参数名称', '')).strip()
                    val = row.get('参数值', None)
                    if pd.isna(val):
                        continue
                    if name == '汇率':
                        df_param.loc[df_param['参数名称'] == name, '参数值'] = exchange_rate
                    elif name == '平台佣金率':
                        df_param.loc[df_param['参数名称'] == name, '参数值'] = platform_commission
                    elif name == '其他费用率':
                        df_param.loc[df_param['参数名称'] == name, '参数值'] = other_fee
                df_param.to_excel(param_path, index=False)
                custom_paths['base_params'] = param_path

            # 保存去年同期订单
            if last_year_file is not None:
                ly_path = os.path.join(base_data_dir, '2025年全年订单底表.xlsx')
                save_uploaded_file(last_year_file, ly_path)

            # Step 2: 加载数据
            progress_bar.progress(20, text="📊 加载数据...")
            loader = DataLoader()
            data = loader.load_all(custom_paths=custom_paths)

            # Step 3: 计算周期 & 去年同期
            progress_bar.progress(40, text="📅 分析数据周期...")
            report_period = calculate_report_period(data)
            last_year_orders = None
            if report_period and last_year_file is not None:
                try:
                    parts = [p.strip() for p in report_period.split('~')]
                    current_start = pd.to_datetime(parts[0])
                    current_end = pd.to_datetime(parts[1])
                    last_year_orders = load_last_year_orders(current_start, current_end, tmpdir)
                except Exception as e:
                    st.warning(f"去年同期数据加载失败: {e}")

            # Step 4: 计算利润
            progress_bar.progress(60, text="🧮 计算利润...")
            formula_file = os.path.join(SYS_DIR, '计算公式详细说明.xlsx')
            calc = ProfitCalculator(
                data,
                last_year_orders=last_year_orders,
                formula_file=formula_file if os.path.exists(formula_file) else None
            )
            result = calc.calculate()

            # Step 5: 生成校验
            progress_bar.progress(80, text="🔍 数据校验...")
            validation = generate_validation_sheets(data, result)

            # Step 6: 生成Excel
            progress_bar.progress(95, text="📄 生成Excel报告...")
            params_info = {
                'exchange_rate': exchange_rate,
                'platform_commission_rate': platform_commission,
                'other_fee_rate': other_fee,
            }
            excel_bytes, timestamp = generate_excel_report(result, validation, report_period, params_info)

            # 保存结果到session_state
            st.session_state.result = result
            st.session_state.result_excel = excel_bytes
            st.session_state.report_period = report_period
            st.session_state.timestamp = timestamp
            st.session_state.calculated = True
            st.session_state.validation = validation

            progress_bar.progress(100, text="✅ 完成！")
            progress_bar.empty()

    except Exception as e:
        progress_bar.empty()
        st.error(f"❌ 测算失败: {str(e)}")
        st.exception(e)

# ═══════════════════════════════════════════════════════════════════════════════
# 结果展示
# ═══════════════════════════════════════════════════════════════════════════════

if st.session_state.calculated and st.session_state.result is not None:
    result = st.session_state.result
    report_period = st.session_state.report_period
    timestamp = st.session_state.timestamp

    st.divider()
    st.success(f"✅ 测算完成！周期: {report_period or '未识别'}")

    # ── 关键指标卡片 ─────────────────────────────────────────────────────────
    st.subheader("📊 关键指标")

    stats = result.summary_stats
    m1, m2, m3, m4, m5 = st.columns(5)
    with m1:
        st.metric("净销售收入", f"${stats.get('总净销售收入', 0):,.2f}")
    with m2:
        st.metric("总毛利润", f"${stats.get('总毛利润', 0):,.2f}")
    with m3:
        gross_margin = stats.get('整体毛利率', 0)
        st.metric("整体毛利率", f"{gross_margin * 100:.2f}%")
    with m4:
        st.metric("总订单数", f"{int(stats.get('总销售数量', 0))}")
    with m5:
        st.metric("总退款金额", f"${stats.get('总退款金额', 0):,.2f}")

    # ── 下载按钮 ─────────────────────────────────────────────────────────────
    st.download_button(
        label="📥 下载利润报告Excel",
        data=st.session_state.result_excel,
        file_name=f"profit_rate_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    # ── 详细数据 ─────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4 = st.tabs(["📋 店铺利润汇总", "📦 SPU利润明细", "📂 三级类目利润明细", "🔍 数据校验"])

    with tab1:
        st.markdown("#### 📋 店铺利润汇总")
        st.markdown(render_store_table(result.store_matrix), unsafe_allow_html=True)

    with tab2:
        render_item_cards(result.spu_matrix, 'SPU', 'SPU')

    with tab3:
        render_item_cards(result.category_matrix, '三级类目', '三级类目')

    with tab4:
        validation = st.session_state.validation
        st.markdown("**SKU成本完整性**")
        st.dataframe(validation['sku_cost'], use_container_width=True, hide_index=True)
        st.markdown("**联盟佣金校验**")
        st.dataframe(validation['affiliate'], use_container_width=True, hide_index=True)
        st.markdown("**广告费校验**")
        st.dataframe(validation['ads'], use_container_width=True, hide_index=True)
        st.markdown("**其他校验**")
        st.dataframe(validation['other'], use_container_width=True, hide_index=True)
